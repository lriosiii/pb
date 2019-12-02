from django.views.generic.base import TemplateView
from .forms import CompareNbRevenueForm, CompareRenewalsRevenueForm, CompareGovgisticsAndRevenueForm, CompareCBAndARForm
import openpyxl
from django.http import HttpResponse
from openpyxl.writer.excel import save_virtual_workbook
import datetime


class AccountingHome(TemplateView):
    template_name = 'accounting/home.html'

    def get_context_data(self, *args, **kwargs):
        context = super().get_context_data(**kwargs)
        context['NBRevenueForm'] = CompareNbRevenueForm
        context['CompareRenewalsRevenueForm'] = CompareRenewalsRevenueForm
        context['CompareGovgisticsAndRevenueForm'] = CompareGovgisticsAndRevenueForm
        context['CompareCBAndARForm'] = CompareCBAndARForm
        return context

    def post(self, request, *args, **kwargs):
        filename = 'myexport'
        from openpyxl.styles import Font, PatternFill
        from openpyxl.styles.colors import RED, BLUE
        try:
            revenue_file_input = request.FILES['reconciliation_file']
            revenue_wb = openpyxl.load_workbook(revenue_file_input, read_only=True, data_only=True)
            rev_accounts = {str(row[5].value).lower() for row in revenue_wb.active.rows if row[5].value}
        except:
            pass

        if request.POST.get('newbiz'):
            resp_wb = openpyxl.load_workbook(request.FILES['nb_file'])
            monthrev = resp_wb.sheetnames[4]
            nb_worksheet = resp_wb[monthrev]
            data_rows = (row for row in nb_worksheet if row[3].value)
            next(data_rows)
            for row in data_rows:
                if str(row[3].value).lower() in rev_accounts:
                    continue
                else:
                    for cell in row:
                        cell.font = Font(color=RED)
        if request.POST.get('renewals'):
            resp_wb = openpyxl.load_workbook(request.FILES['renewals_file'])
            sheet12 = resp_wb.sheetnames[11]
            worksheet = resp_wb[sheet12]
            data_rows = (row for row in worksheet if row[2].value)
            next(data_rows)
            for row in data_rows:
                if str(row[2].value).lower() in rev_accounts:
                    continue
                else:
                    for cell in row:
                        cell.font = Font(color=RED)

        if request.POST.get('govgistics'):
            resp_wb = openpyxl.load_workbook(request.FILES['govgistics_file'])
            sheet3 = resp_wb.sheetnames[2]
            worksheet = resp_wb[sheet3]
            data_rows = (row for row in worksheet if row[6].value)
            next(data_rows)
            for row in data_rows:
                if str(row[6].value).lower() in rev_accounts:
                    continue
                else:
                    for cell in row[4:]:
                        cell.font = Font(color=RED)

        if request.POST.get('cbar'):
            filename, date = 'AGEING-CB-PR', datetime.datetime.now().date()
            cb_wb = openpyxl.load_workbook(request.FILES['CB_file'], read_only=True)
            resp_wb = openpyxl.load_workbook(request.FILES['AR_file'])
            cb_worksheet = cb_wb.active
            ws = resp_wb.active
            cb_all, cb_orngyellow_invoices, cb_orngfill, cb_bluetext = list(), list(), list(), list()
            cb_filteredrows = (row for row in cb_worksheet.iter_rows(min_row=3) if row[0].value is not None and int(row[0].value) >= 45)
            for row in cb_filteredrows:
                cb_all.append(row[5].value)         # yellow     # orange
                if row[0].fill.start_color.index in ('FFFFFF00', 'FFFFC000'):
                    cb_orngyellow_invoices.append(row[5].value)         # blue
                elif row[0].value >= 65 and getattr(row[0].font.color, 'value', 'default') in ('FF00B0F0'):
                    cb_orngfill.append(row[5].value)
                elif row[0].value < 65 and getattr(row[0].font.color, 'value', 'default') in ('FF00B0F0'):
                    cb_bluetext.append(row[5].value)
            ws_populatedrows = (row for row in ws.iter_rows(min_row=3) if row[0].value is not None)
            for row in ws_populatedrows:
                if int(row[0].value) < 45:
                    ws.row_dimensions[row[0].row].hidden = True
                if row[5].value in cb_orngyellow_invoices:
                    for cell in row:
                        cell.fill = PatternFill("solid", fgColor="FFFFFF00")
                elif row[5].value in cb_orngfill:
                    for cell in row:
                        cell.fill = PatternFill("solid", fgColor="FFFFC000")
                elif row[5].value in cb_bluetext:
                    for cell in row:
                        cell.font = Font(color=BLUE)
                elif row[0].value >= 45 and row[5].value not in cb_all:
                    for cell in row:
                        cell.fill = PatternFill("solid", fgColor="FFFFC000")

        response = HttpResponse(content=save_virtual_workbook(resp_wb), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={filename}_{date}.xlsx'
        return response

